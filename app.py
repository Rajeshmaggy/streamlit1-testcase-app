import streamlit as st
import pandas as pd
import os
# import cv2
import pytesseract
from fpdf import FPDF  # Import the FPDF library
import tempfile
import groq
# import os
import re
import PyPDF2
from PIL import Image, ImageDraw, ImageFont
import pytesseract
from PyPDF2 import PdfReader
from pdfminer.high_level import extract_text
from docx import Document
import fitz  # PyMuPDF
import torch
from transformers import AutoModel, AutoTokenizer
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.pipeline import Pipeline
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score, f1_score, classification_report
from sklearn.utils.class_weight import compute_class_weight
from transformers import BertForSequenceClassification, AutoTokenizer
from transformers import BertTokenizer, BertForSequenceClassification
from torch.utils.data import DataLoader, TensorDataset, random_split
from sklearn.preprocessing import LabelEncoder
import numpy as np
from groq import Groq
from torch.optim import AdamW
import torch.nn.functional as F
# from warnings import filterwarnings
import os
import torch
from torch.utils.data import DataLoader, TensorDataset
from transformers import BertTokenizer, BertForSequenceClassification, AdamW
from sklearn.preprocessing import LabelEncoder
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from transformers import pipeline

# Required Libraries for Document Handling
from pdfminer.high_level import extract_text as extract_text_from_pdf
from docx import Document
from pytesseract import image_to_string
from PIL import Image
import openpyxl
from openpyxl import load_workbook

# Required Library for the Groq API
# from groq.api import Groq

# Set up pytesseract path for OCR
pytesseract.pytesseract.tesseract_cmd = "/usr/bin/tesseract"

# Initialize Groq LLaMA3 Client
client = groq.Client(api_key='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')  # Replace with your actual Groq API key

# Set up the page
st.set_page_config(page_title="Test Case Generator", layout="wide")

# File path to store data
DATA_FILE = "test_cases.csv"
USER_CREDENTIALS_FILE = "user_credentials.csv"

# Load existing data or create a new DataFrame
if os.path.exists(DATA_FILE):
    test_cases_df = pd.read_csv(DATA_FILE)
else:
    test_cases_df = pd.DataFrame(columns=["Email", "Test Case Type", "File Name"])

# Load user credentials from the file
if os.path.exists(USER_CREDENTIALS_FILE):
    user_credentials_df = pd.read_csv(USER_CREDENTIALS_FILE)
else:
    user_credentials_df = pd.DataFrame(columns=["Email", "Password"])

# Function to save user credentials
def save_user_credentials(name,email, password):
    global user_credentials_df
    new_entry = pd.DataFrame({"Name": [name],"Email": [email], "Password": [password]})
    user_credentials_df = pd.concat([user_credentials_df, new_entry], ignore_index=True)
    user_credentials_df.to_csv(USER_CREDENTIALS_FILE, index=False)

# Function to check if credentials are valid
def check_credentials(email, password):
    user = user_credentials_df[user_credentials_df["Email"] == email]
    if not user.empty and user["Password"].values[0] == password:
        return True
    return False

# Function to save test case data
def save_data(email, test_case_type, file_name):
    global test_cases_df
    new_entry = pd.DataFrame(
        {"Email": [email], "Test Case Type": [test_case_type], "File Name": [file_name]}
    )
    test_cases_df = pd.concat([test_cases_df, new_entry], ignore_index=True)
    test_cases_df.to_csv(DATA_FILE, index=False)

# Function to extract frames from video
def extract_frames(video_path, interval=1):
    cap = cv2.VideoCapture(video_path)
    frames = []
    frame_count = 0
    success = True

    while success:
        success, frame = cap.read()
        if frame_count % (interval * 30) == 0:  # Capture frame every `interval` seconds
            if success:
                frames.append(frame)
        frame_count += 1

    cap.release()
    return frames

# Function to extract text from a frame using OCR
def extract_text_from_frame(frame):
    gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    text = pytesseract.image_to_string(gray_frame)
    return text

# Function to prepare the input for LLaMA3 API
def prepare_input_for_llama(extracted_text):
    prompt = f"""
    The following text has been extracted from screens of an app. Please generate a Product Requirements Document (PRD) with sections like Overview, Objectives, Functional Requirements, and Non-Functional Requirements based on the app screens described below:

    Extracted Text:
    {extracted_text}

    PRD:
    """
    return prompt

# Function to generate PRD using Groq LLaMA3 API
def generate_prd_from_text(text):
    prompt = prepare_input_for_llama(text)

    # API call to LLaMA 3 using Groq client
    chat_completion = client.chat.completions.create(
        messages=[{
            "role": "user",
            "content": prompt,
        }],
        model="llama3-70b-8192",  # Use the appropriate model
        temperature=0  # Set temperature to 0 for deterministic output
    )

    # Access the generated PRD content
    response_message = chat_completion.choices[0].message.content.strip()

    return response_message
#for uploaded pdf reader  
def read_pdf_document(file_path):
    """Reads a PDF document and returns the text content."""
    full_text = []
    try:
        with fitz.open(file_path) as doc:
            for page in doc:
                full_text.append(page.get_text())
        return '\n'.join(full_text)
    except Exception as e:
        print(f"Error reading PDF file: {e}")
        return ""
# for storing the txt in variable
def read_pdf_to_variable(pdf_path):
            with open(pdf_path, "rb") as file:
                pdf_data = file.read()
            return pdf_data
# img to txt
def image_to_text(file_path):
    """Extracts text from an image using OCR and returns the extracted text."""
    try:
        with Image.open(file_path) as img:
            extracted_text = pytesseract.image_to_string(img)
            # st.write(extracted_text)
        return extracted_text
    except Exception as e:
        print(f"An error occurred: {e}")
        return ""
# Function to save PRD document to PDF
def save_text_to_pdf(text, output_pdf_path):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Arial", size=12)

    # Split the text into lines and add each line to the PDF
    for line in text.split('\n'):
        pdf.multi_cell(0, 10, line)

    # Save the PDF to the specified file path
    pdf.output(output_pdf_path)

# Main section content
if 'login_successful' not in st.session_state or not st.session_state.login_successful:
    # Login / Signup page content
    st.markdown("<h1 style='text-align: center;'>Welcome to the Test Case Generator</h1>", unsafe_allow_html=True)
    

    login_or_signup = st.radio("Choose action:", ["Login", "Sign Up"], index=0)

    if login_or_signup == "Login":
        
        user_email = st.text_input("Email")
        user_password = st.text_input("Password", type="password")
        login_button = st.button("Login")

        if login_button:
            if check_credentials(user_email, user_password):
                st.session_state.login_successful = True
                st.session_state.user_email = user_email
                st.success("Login successful!")
            else:
                st.error("Invalid credentials. Please try again.")

    elif login_or_signup == "Sign Up":
        user_name = st.text_input("Name")
        new_user_email = st.text_input("Email")
        new_user_password = st.text_input("Password", type="password")
        confirm_password = st.text_input("Confirm Password", type="password")
        signup_button = st.button("Sign Up")

        if signup_button:
            if new_user_password != confirm_password:
                st.error("Passwords do not match. Please try again.")
            elif new_user_email in user_credentials_df["Email"].values:
                st.error("Email already exists. Please log in.")
            else:
                save_user_credentials(user_name,new_user_email, new_user_password)
                st.success("Signup successful! You can now log in.")
else:
    # Display the profile icon and user info in the top-right corner
    st.markdown("""
    <style>
        .profile-container {
            display: flex;
            justify-content: flex-end;
            align-items: center;
            position: absolute;
            top: -40px;
            right: -35px;
        }
        .profile-icon {
            width: 30px;
            height: 30px;
            border-radius: 50%;
            background-color: #0073e6;
            color: white;
            display: flex;
            justify-content: center;
            align-items: center;
            font-size: 18px;
        }
        .profile-info {
            margin-left: 10px;
            font-size: 14px;
        }
    </style>
    """, unsafe_allow_html=True)

# Profile icon (using the first letter of the email as the icon)
    profile_icon = st.session_state.user_email[0].upper()  # Using the first letter of the email as the icon
    st.markdown(f"""
    <div class="profile-container">
        <div class="profile-icon">{profile_icon}</div>
        <div class="profile-info">{st.session_state.user_email.split("@")[0]}</div>
    </div>
""", unsafe_allow_html=True)

    # Sidebar - User History
    with st.sidebar:
        # uploaded_file = st.file_uploader("Upload File", type=["mp4", "mov", "avi", "pdf", "docx", "png", "jpg", "jpeg"])
        st.markdown("### Test Case History")
        user_history = test_cases_df[test_cases_df["Email"] == st.session_state.user_email]
        if user_history.empty:
            st.info("No test cases submitted yet.")
        else:
            for _, row in user_history.iterrows():
                st.markdown(f"- **{row['Test Case Type']}**: {row['File Name']}")

    # Main Section Content
    st.markdown("<h1 style='text-align: center;'>Test Case Generator</h1>", unsafe_allow_html=True)

    # Test Case Input Section
    col1, col2, col3 = st.columns([1, 1, 1])

    with col1:
        test_case_type = st.selectbox(
            "Test Case Type",
            ["Test Case Generation", "Test Case Validation", "Context Modeling"]
        )

    with col2:
        content_type = st.selectbox(
            "Content Type",
            ["Screenshots", "Video", "Document","xl"]
        )
    with col3:

# File uploader inside the column (col3)
        uploaded_file = col3.file_uploader("Upload File", type=["mp4", "mov", "avi", "pdf", "docx", "png", "jpg", "jpeg","csv"])

# Check if a file is uploaded
# if uploaded_file:
#     # Determine the content type based on the uploaded file's extension
#     file_extension = uploaded_file.name.split(".")[-1].lower()

#     if file_extension in ["mp4", "mov", "avi"]:
#         content_type = "Video"
#     elif file_extension in ["pdf", "docx"]:
#         content_type = "Document"
#     elif file_extension in ["png", "jpg", "jpeg"]:
#         content_type = "Photo"
#     else:
#         content_type = "Unknown"

#     # Display the content type of the uploaded file
#     st.write(f"The uploaded file is a {content_type}.")


    if st.button("Submit Test Case"):
        if not uploaded_file:
            st.warning("Please upload a file.")
        elif test_case_type == "Test Case Calidation":
            with tempfile.NamedTemporaryFile(delete=False) as temp_file:
                temp_file.write(uploaded_file.read())
                video_path = temp_file.name
            try:
                    # Load the Excel file
                    wb = load_workbook(video_path)
                    ws = wb.active
                    df = pd.read_excel(video_path)

                    # Validate columns
                    if "Expected Results" not in df.columns:
                        st.error("Missing 'Expected Results' column.")
                        st.error(f"File {uploaded_file.name} is missing the 'Expected Results' column.")
                        
                        # continue

                    # Select columns up to "Expected Results"
                    idx = df.columns.get_loc("Expected Results")
                    df_selected = df.iloc[:, :idx + 1]

                    # Filter out rows with strikethrough in any cell
                    rows_to_keep = []
                    for row_idx, row in df.iterrows():
                        keep_row = True
                        for col_idx in range(len(df.columns)):
                            cell = ws.cell(row=row_idx + 2, column=col_idx + 1)
                            if cell.font and cell.font.strike:
                                keep_row = False
                                break
                        if keep_row:
                            rows_to_keep.append(row_idx)
                    df_filtered = df.iloc[rows_to_keep]

                    # Perform test case evaluation using Groq API
                    results = {
                        "Test Case Number": [],
                        "Clarity": [],
                        "Coverage": [],
                        "Overall Quality": []
                    }

                    for idx, test_case in df_filtered.iterrows():
                        scenario = test_case["Scenario"]
                        prerequisite = test_case.get("Prerequisite", "")
                        test_steps = test_case["Test Steps"]
                        expected_result = test_case["Expected Results"]

                        # Construct prompt for Groq
                        prompt = f"""
                        You are an expert in QA. Evaluate this test case and provide:
                        1. Scores out of 10 for: Clarity, Coverage, and Overall Quality.
                        2. Specific suggestions to improve each category.

                        Scenario: {scenario}
                        Precondition: {prerequisite}
                        Test Steps: {test_steps}
                        Expected Result: {expected_result}
                        Respond with scores and recommendations for improvement.
                        """

                        try:
                            response = client.chat.completions.create(
                                messages=[{"role": "user", "content": prompt}],
                                model="llama3-70b-8192",
                                temperature=0
                            )
                            evaluation = response.choices[0].message.content

                            # Extract scores and suggestions from the response
                            clarity_match = re.search(r"1\.\s*Clarity:\s*(\d+)/10", evaluation)
                            coverage_match = re.search(r"2\.\s*Coverage:\s*(\d+)/10", evaluation)
                            overall_quality_match = re.search(r"3\.\s*Overall Quality:\s*(\d+)/10", evaluation)

                            clarity = clarity_match.group(1) if clarity_match else "N/A"
                            coverage = coverage_match.group(1) if coverage_match else "N/A"
                            overall_quality = overall_quality_match.group(1) if overall_quality_match else "N/A"

                            results["Test Case Number"].append(f"TC {idx + 1}")
                            results["Clarity"].append(f"{clarity}/10")
                            results["Coverage"].append(f"{coverage}/10")
                            results["Overall Quality"].append(f"{overall_quality}/10")

                        except Exception as e:
                            st.error(f"Groq API error for test case {idx + 1}: {e}")
                            results["Test Case Number"].append(f"TC {idx + 1}")
                            results["Clarity"].append("Error")
                            results["Coverage"].append("Error")
                            results["Overall Quality"].append("Error")

                    # Merge results into the original DataFrame
                    df_result = pd.DataFrame(results)
                    df_merged = pd.merge(df, df_result, left_index=True, right_index=True, how="left")

                    # Save the results to a new Excel file
                    output_file = f"validated_{uploaded_file.name}"
                    df_merged.to_excel(output_file, index=False)
                    st.success(f"Results saved to {output_file}")

                    # Allow download of the validated file
                    with open(output_file, "rb") as f:
                        st.download_button(
                            label="Download Validation Results",
                            data=f,
                            file_name=output_file,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )

            except Exception as e:
                st.error(f"Error processing file {uploaded_file.name}: {e}")
        
        elif test_case_type == "Test Case Generation" and content_type == "Video":
            # output_pdf_path = None
            with tempfile.NamedTemporaryFile(delete=False) as temp_file:
                temp_file.write(uploaded_file.read())
                video_path = temp_file.name

        # Step 1: Extract frames from the video
            frames = extract_frames(video_path, interval=2)

        # Step 2: Extract text from each frame using OCR
            extracted_text = ""
            for idx, frame in enumerate(frames):
                text = extract_text_from_frame(frame)
                if text.strip():  # If there's any text detected, add it
                    extracted_text += f"\n--- Frame {idx} ---\n{text}\n"

        # Step 3: Generate PRD from extracted text using Groq LLaMA3
            if extracted_text.strip():
                prd_document = generate_prd_from_text(extracted_text)
                st.success("PRD document generated successfully!")

        # Path for the document
            output_pdf_path = "prd_document.pdf"
            # prd_file_path = output_pdf_path

        # Step 4: Save the generated PRD document to PDF
        # Assuming this code is after the PRD generation step
            a = save_text_to_pdf(prd_document, output_pdf_path)

        # Allow the user to download the PRD PDF
            def read_pdf_to_variable(pdf_path):
                with open(pdf_path, "rb") as file:
                    pdf_data = file.read()
                return pdf_data

            pdf_data = read_pdf_to_variable(output_pdf_path)
            st.download_button(
            label="Download PRD PDF",
            data=pdf_data,
            file_name="prd_document.pdf",
            mime="application/pdf"
        )

        else:
            output_pdf_path = None
        # If the content type is not "Video" (e.g., for "Photo" or "Document")
            with tempfile.NamedTemporaryFile(delete=False) as temp_file:
                temp_file.write(uploaded_file.read())
                uploaded_file_path = temp_file.name

        # Handle the file for non-video content types (Photo, Document, etc.)
            if content_type == "Screenshots":
            # Perform actions specific to photos (e.g., processing or OCR if needed)
                extracted_text = image_to_text(uploaded_file_path)
                if extracted_text.strip():
                    prd_document = generate_prd_from_text(extracted_text)
                    # st.success("PRD document generated successfully!")
                output_pdf_path = "prd_document.pdf"
                prd_file_path = output_pdf_path
                save_text_to_pdf(prd_document, output_pdf_path)
            # st.image(uploaded_file_path, caption="Uploaded Photo")
        
            elif content_type == "Document":
            # If it's a document, extract text or do any processing
                extracted_text = read_pdf_document(uploaded_file_path)
                if extracted_text.strip():
                    prd_document = generate_prd_from_text(extracted_text)
                    st.success("PRD document generated successfully!")
            
            # Save the generated PRD document to PDF
                output_pdf_path = "prd_document.pdf"
                prd_file_path = output_pdf_path
                save_text_to_pdf(prd_document, output_pdf_path)

            # Allow the user to download the PRD PDF
                pdf_data = read_pdf_to_variable(output_pdf_path)
            # st.download_button(
            #     label="Download PRD PDF",
            #     data=pdf_data,
            #     file_name="prd_document.pdf",
            #     mime="application/pdf"
            # )
        
        # output_pdf_path = prd_document
        ############################################
        
        # Filter warnings
        
        # filterwarnings("ignore")
        
        # Define the models for each PRD domain
        model_mapping = {
            'e-commerce': 'EZlee/e-commerce-bert-base-multilingual-cased',
            'financial': 'ahmedrachid/FinancialBERT',
            'biological': 'dmis-lab/biobert-v1.1',
            'common': 'bert-large-uncased'
        }
        
        def read_word_document(file_path):
            """Reads a Word document and returns the text content."""
            doc = Document(file_path)
            full_text = [para.text for para in doc.paragraphs if para.text.strip() != '']
            return '\n'.join(full_text)
        
        def read_pdf_document(file_path):
            """Reads a PDF document and returns the text content."""
            full_text = []
            try:
                with fitz.open(file_path) as doc:
                    for page in doc:
                        full_text.append(page.get_text())
                return '\n'.join(full_text)
            except Exception as e:
                print(f"Error reading PDF file: {e}")
                return ""
        
        def sanitize_text(text):
            """Sanitizes text to remove XML-incompatible characters."""
            sanitized_text = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', text)
            return sanitized_text
        
        
        def preprocess_text(text):
            """Preprocess the text without altering any characters, symbols, or formatting."""
            return sanitize_text(text)
        
        def image_to_text(file_path):
            """Extracts text from an image using OCR and returns the extracted text."""
            try:
                with Image.open(file_path) as img:
                    extracted_text = pytesseract.image_to_string(img)
                    # st.write(extracted_text)
                return extracted_text
            except Exception as e:
                print(f"An error occurred: {e}")
                return ""
        
        
        def train_bert_classifier(X, y, model_name='bert-base-uncased', max_length=256, batch_size=32, epochs=4, lr=3e-5):
            """Trains a BERT classifier on the given text data with improvements for better performance."""
        
            # Check if CUDA (GPU) is available, otherwise use CPU
            device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
            print(f"Using device: {device}")
        
            # Load tokenizer and model
            tokenizer = BertTokenizer.from_pretrained(model_name)
            label_encoder = LabelEncoder()
            y_encoded = label_encoder.fit_transform(y)
            num_labels = len(label_encoder.classes_)
        
            model = BertForSequenceClassification.from_pretrained(model_name, num_labels=num_labels)
        
            # Move model to device (e.g., GPU if available)
            model = model.to(device)
        
            # Tokenize input texts
            inputs = tokenizer(X, padding=True, truncation=True, max_length=max_length, return_tensors="pt")
            dataset = TensorDataset(inputs['input_ids'], inputs['attention_mask'], torch.tensor(y_encoded))
        
            # Split dataset into train and validation sets
            train_size = int(0.8 * len(dataset))
            val_size = len(dataset) - train_size
            train_dataset, val_dataset = random_split(dataset, [train_size, val_size])
        
            train_loader = DataLoader(train_dataset, batch_size=batch_size, shuffle=True)
            val_loader = DataLoader(val_dataset, batch_size=batch_size)
        
            # Compute class weights to handle class imbalance
            class_weights = compute_class_weight('balanced', classes=np.unique(y_encoded), y=y_encoded)
            class_weights = torch.tensor(class_weights, dtype=torch.float).to(device)
        
            optimizer = AdamW(model.parameters(), lr=lr)
        
            # Training loop
            model.train()
            for epoch in range(epochs):
                total_loss = 0
                for batch in train_loader:
                    input_ids, attention_mask, labels = [b.to(device) for b in batch]
                    optimizer.zero_grad()
                    outputs = model(input_ids=input_ids, attention_mask=attention_mask, labels=labels)
                    loss = F.cross_entropy(outputs.logits, labels, weight=class_weights)
                    loss.backward()
                    optimizer.step()
                    total_loss += loss.item()
        
                print(f"Epoch {epoch + 1}, Loss: {total_loss / len(train_loader)}")
        
            # Evaluate the model on the validation set
            model.eval()
            all_preds = []
            all_labels = []
            with torch.no_grad():
                for batch in val_loader:
                    input_ids, attention_mask, labels = [b.to(device) for b in batch]
                    outputs = model(input_ids=input_ids, attention_mask=attention_mask)
                    preds = torch.argmax(outputs.logits, dim=1)
                    all_preds.extend(preds.cpu().tolist())
                    all_labels.extend(labels.cpu().tolist())
        
            # Convert predicted indices and true labels back to original labels
            all_preds_labels = label_encoder.inverse_transform(all_preds)
            all_labels_true = label_encoder.inverse_transform(all_labels)
        
            # Calculate accuracy and generate classification report
            accuracy = accuracy_score(all_labels_true, all_preds_labels)
            f1 = f1_score(all_labels_true, all_preds_labels, average='weighted')
        
            print(f"Validation Accuracy: {accuracy:.4f}")
            print(f"F1 Score: {f1:.4f}")
            #print("Classification Report:\n", classification_report(all_labels_true, all_preds_labels, labels=label_encoder.classes_, target_names=label_encoder.classes_))
        
            print("Model fine-tuned successfully!")
            return model, tokenizer, label_encoder
        
        
        def select_model(prd_text, classifier, tokenizer, label_encoder, max_length=512):
            """Selects the appropriate BERT model based on the PRD domain classification."""
            inputs = tokenizer(prd_text, padding=True, truncation=True, max_length=max_length, return_tensors="pt")
            with torch.no_grad():
                outputs = classifier(**inputs)
            domain_index = torch.argmax(outputs.logits, dim=1).item()
            domain = label_encoder.inverse_transform([domain_index])[0]
        
            if domain in model_mapping:
                model_name = model_mapping[domain]
                tokenizer = AutoTokenizer.from_pretrained(model_name)
                model = BertForSequenceClassification.from_pretrained(model_name)
                print(f"Selected Model: {model_name}")
            else:
                print("Domain not found in model mapping, defaulting to 'common'")
                model_name = model_mapping['common']
                tokenizer = AutoTokenizer.from_pretrained(model_name)
                model = BertForSequenceClassification.from_pretrained(model_name)
        
            return model, tokenizer
        
        def refine_prd_with_model(prd_text, model, tokenizer, max_length=512):
            """Refines the PRD text using the selected BERT model."""
        
            # Tokenize the input text
            inputs = tokenizer(prd_text, return_tensors='pt', truncation=True, padding=True, max_length=max_length)
        
            # Move the input tensors to the correct device (same as the model)
            device = next(model.parameters()).device
            inputs = {key: val.to(device) for key, val in inputs.items()}
        
            # Ensure no token exceeds the model's vocabulary size
            vocab_size = model.config.vocab_size
            input_ids = inputs['input_ids']
        
            # Replace any token indices that exceed the vocabulary size with the [UNK] token index
            unk_token_index = tokenizer.unk_token_id  # Get the index for the [UNK] token
            input_ids[input_ids >= vocab_size] = unk_token_index
        
            # Forward pass with no gradient calculation
            with torch.no_grad():
                outputs = model(**inputs)
        
            # Placeholder for refining PRD logic - you can implement specific refinement here
            refined_prd = prd_text  # Modify this to perform actual text refinement if needed.
        
            return refined_prd
        
        
        def generate_test_cases(prd_text, api_key):
            """Generates test cases from the refined PRD using the Groq API."""
            client = Groq(api_key=api_key)
        
            prompt = f"""
            Based on the following PRD document content, generate comprehensive positive and negative test cases that cover all requirements. Ensure that each test case includes:
        
            1. Precondition
            2. Test steps
            3. Expected result
            4. Requirement ID (if applicable) or reference to the specific section of the PRD
            5. Whether it is a positive or negative test case
        
            PRD Content:
            {prd_text}
        
            Please ensure that the test cases are aligned with industry standards and cover all functional and non-functional requirements, including edge cases and failure scenarios.
            and give me the result in table format
            
           
            """
        
            completion = client.chat.completions.create(
                messages=[
                    {"role": "system", "content": "You are an expert test case generator specializing in PRD analysis."},
                    {"role": "user", "content": prompt}
                ],
                model="llama3-70b-8192",
                max_tokens=3000,
                temperature=0,
                top_p=0.9,
            )
        
            generated_test_cases = completion.choices[0].message.content.strip()
            return generated_test_cases
        
        def fine_tune_model_with_prd(prd_text, model, tokenizer, epochs=2, max_length=512):
            """Fine-tunes the selected model using the input PRD text."""
            # Tokenize the PRD text and create a dataset
            inputs = tokenizer(prd_text, padding=True, truncation=True, max_length=max_length, return_tensors="pt")
            labels = torch.tensor([1])  # Assuming binary label, adjust based on the task
        
            dataset = TensorDataset(inputs['input_ids'], inputs['attention_mask'], labels)
            train_loader = DataLoader(dataset, batch_size=1, shuffle=True)
        
            optimizer = AdamW(model.parameters(), lr=5e-5)
            model.train()
        
            # Fine-tuning loop
            for epoch in range(epochs):
                total_loss = 0
                for batch in train_loader:
                    input_ids, attention_mask, labels = batch
                    optimizer.zero_grad()
                    outputs = model(input_ids=input_ids, attention_mask=attention_mask, labels=labels)
                    loss = outputs.loss
                    loss.backward()
                    optimizer.step()
                    total_loss += loss.item()
        
                print(f"Fine-tuning Epoch {epoch + 1}, Loss: {total_loss / len(train_loader)}")
        
            print("Model fine-tuned with PRD data!")
            return model
        
        prd_file_path = output_pdf_path  # Update this path
        api_key = "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"
        
        # Read PRD file based on its extension
        if prd_file_path.endswith('.docx'):
            prd_text = read_word_document(prd_file_path)
            st.success("PRD document doc read successfully!")
            output_in_image = False
        elif prd_file_path.endswith('.pdf'):
            prd_text = read_pdf_document(prd_file_path)
            st.success("PRD document pdf read successfully!")
            # st.success("Hey im training the bert!")
            output_in_image = False
        elif prd_file_path.endswith(('.png', '.jpg', '.jpeg')):
            prd_text = image_to_text(prd_file_path)
            st.success("PRD document img read successfully!")
            output_in_image = True
        else:
            print("Unsupported file format.")
            prd_text = ""
            # st.success("PRD document text read successfully!")
            output_in_image = False
        
        if prd_text:
            # Preprocess the PRD text
            prd_text = preprocess_text(prd_text)
            # st.success("PRD document pdf read successfully!")
        
        # Sample training data for domain classification
        X_train = [
            # E-commerce examples
            "Product details include name, description, category, specifications, and images.",
            "Pricing information like cost, discounts, taxes, and currency details are required.",
            "Manage inventory with stock levels, SKU, and location tracking.",
            "User interface elements such as navigation, filters, and sorting are important for usability.",
            "Order management involves order creation, tracking, status updates, returns, and cancellations.",
            "Payment and checkout processes cover various payment methods, gateways, and security protocols.",
            "User profiles include account creation, login features, preferences, and order history.",
            "Shipping options, delivery tracking, estimated delivery time, and address management are included.",
            "Notifications for orders, promotions, and updates are sent via emails, SMS, and in-app alerts.",
            "Analytics and reporting track user behavior, sales performance, and customer feedback.",
            "Product recommendations based on user behavior and purchase history.",
            "Handling customer reviews, ratings, and feedback for products.",
            "Integration with social media platforms for product sharing and promotions.",
            "Wishlists and saved items functionality for users.",
            "Product comparison tools for evaluating multiple items.",
            "Cross-selling and up-selling techniques integrated into the platform.",
            "Gift cards, vouchers, and promotional codes management.",
            "Customer loyalty programs and reward points systems.",
            "Affiliate marketing and partner program integration.",
            "Return and refund policies, along with exchange procedures.",
            "Multiple currency support for international customers.",
            "Mobile-responsive design and features for shopping on smartphones.",
            "Product search functionality with advanced filtering and sorting options.",
            "Handling flash sales, limited-time offers, and special deals.",
            "SEO optimization for product pages and category listings.",
            "Handling user-generated content like reviews, photos, and videos.",
            "Fraud detection and prevention measures for transactions.",
            "Inventory restocking alerts and notifications for low-stock items.",
            "Bulk order processing and wholesale management.",
            "Handling multiple payment gateways and international payment options.",
            "Customer service features like live chat and helpdesk integration.",
            "Multilingual support for a global audience.",
            "Personalized marketing campaigns based on user data.",
            "Order history and invoice management for users.",
            "Product bundles and combo offers.",
            "Abandoned cart recovery features.",
            "Gift-wrapping options and special instructions for delivery.",
            "Product lifecycle management and inventory forecasting.",
            "Handling digital products like eBooks, software, and licenses.",
            "Integration with courier services for automated shipping updates.",
            "Managing different product variants like sizes, colors, and styles.",
            "Vendor management and supplier integration.",
            "Customer satisfaction surveys and feedback collection.",
            "User-friendly checkout process with minimal steps.",
            "Handling returns logistics and reverse shipping.",
            "Dynamic pricing strategies based on demand and competition.",
            "Creating and managing customer segments for targeted marketing.",
            "Data security and privacy compliance for user information.",
            "Handling seasonal sales events like Black Friday, Cyber Monday.",
            "Automatic product updates from suppliers and manufacturers.",
            "Product_Categorization: Organizing products into categories and subcategories.",
            "Customer_Support: Handling customer queries, complaints, and support tickets.",
            "Order_Fulfillment: Processing, packaging, and shipping orders to customers.",
            "Payment_Refunds: Managing refunds and reimbursements for returned products.",
            "Search_Engine_Optimization: Improving product page visibility in search engines.",
            "Marketplace_Integration: Integrating with external marketplaces like Amazon, eBay.",
            "Subscription_Management: Handling recurring payments and subscription-based services.",
            "Supplier_Management: Coordinating with suppliers for product availability and quality.",
            "Digital_Marketing: Running email campaigns, PPC ads, and social media promotions.",
            "Multi_Vendor_Platform: Managing multiple sellers on the same platform.",
        
            # Financial examples
            "Account Management: Account types, creation, modification, closure.",
            "Transactions: Types of transactions (deposits, withdrawals, transfers), transaction limits.",
            "Security: Authentication, authorization, fraud detection, encryption.",
            "Compliance and Regulation: KYC (Know Your Customer), AML (Anti-Money Laundering), GDPR.",
            "Reporting and Analytics: Financial reporting, risk analysis, performance metrics.",
            "Payment Processing: Payment methods, processing time, charges, reconciliation.",
            "Investment Products: Types (stocks, bonds, mutual funds), portfolios, market data integration.",
            "Alerts and Notifications: Transaction alerts, account changes, policy updates.",
            "Customer Support: Chatbots, FAQs, helpdesk functionality, grievance handling.",
            "API Integration: Third-party services, data exchange protocols.",
            "Market trends, investment strategies, and economic analysis.",
            "Guidelines for financial compliance, risk management, and auditing processes.",
            "Financial forecasts, budgets, and expenditure tracking methodologies.",
            "Banking operations: Loans, credit lines, mortgage management.",
            "Taxation: Income tax, corporate tax, VAT, tax returns.",
            "Insurance: Policy management, claims processing, premium calculations.",
            "Loan origination and approval process, including credit scoring.",
            "Wealth management: Asset allocation, portfolio management, financial advising.",
            "Digital wallets and cryptocurrency transactions.",
            "Pension and retirement plans, fund management, disbursements.",
            "Credit card processing, rewards programs, and interest calculations.",
            "Merchant services: Payment gateway integration, POS systems.",
            "Forex trading: Currency pairs, exchange rates, and risk management.",
            "Investment banking: Mergers, acquisitions, capital raising, underwriting.",
            "Personal finance management tools: Budgeting, saving, investing.",
            "Risk management: Hedging strategies, derivatives, insurance.",
            "Financial audits: Internal controls, compliance checks, financial statements.",
            "Credit scoring and risk assessment for loans and credit cards.",
            "Anti-fraud measures for online banking and payment systems.",
            "Corporate finance: Capital structure, dividend policy, financing decisions.",
            "Liquidity management: Cash flow analysis, short-term investments.",
            "Regulatory reporting for financial institutions.",
            "Transaction reconciliation and settlement processes.",
            "Wealth transfer planning: Inheritance, trusts, estate planning.",
            "Corporate governance and ethics in finance.",
            "Financial literacy and education programs.",
            "Microfinance and financial inclusion initiatives.",
            "Behavioral finance: Investor psychology, market anomalies.",
            "Crowdfunding platforms and alternative financing models.",
            "Venture capital and private equity: Fundraising, valuation, exits.",
            "Fintech innovation: Blockchain, robo-advisors, AI in finance.",
            "International finance: Global markets, exchange rate risk, trade finance.",
            "Capital markets: Stocks, bonds, derivatives trading.",
            "Financial modeling and valuation techniques.",
            "Mergers and acquisitions: Due diligence, integration planning.",
            "Corporate restructuring and turnaround management.",
            "Asset-backed securities and securitization.",
            "Real estate finance: Mortgage-backed securities, REITs.",
            "Green finance: Sustainable investing, carbon credits, ESG.",
            "Trade finance: Letters of credit, export financing, trade insurance.",
            "Credit_Risk_Assessment: Evaluating the creditworthiness of individuals and businesses.",
            "Debt_Management: Handling loans, repayments, and debt collection.",
            "Financial_Planning: Creating financial strategies and investment plans for clients.",
            "Derivatives_Trading: Trading financial contracts like options and futures.",
            "Wealth_Advisory: Offering personalized investment advice and financial counseling.",
            "Hedge_Funds: Managing alternative investment vehicles for high-net-worth individuals.",
            "Pension_Funds: Managing retirement savings plans and pension investments.",
            "Asset_Liability_Management: Managing assets and liabilities to ensure financial stability.",
            "Trade_Finance: Financing international trade and managing trade-related risks.",
            "Corporate_Treasury: Managing a corporation's liquidity, investments, and risk.",
            # Biological examples
            "Biotech PRD covers gene therapy innovations, new pharmaceutical developments, and clinical trials.",
            "Details on laboratory techniques, biological markers, and sample preparation are provided.",
            "The document includes sections on ethical considerations and regulatory compliance for biotech.",
            "Research Objectives: Purpose, hypothesis, expected outcomes, study design.",
            "Experimental Protocols: Methodologies, equipment, materials, control groups.",
            "Data Collection: Data types (genomic, proteomic, etc.), collection methods, storage format.",
            "Compliance and Ethics: Ethical approval, patient consent, data privacy (HIPAA, GDPR).",
            "Statistical Analysis: Models, significance levels, software/tools used.",
            "Results and Findings: Data visualization, results interpretation, conclusions.",
            "Collaboration and Sharing: Data sharing policies, collaborative tools, access controls.",
            "Publication and Documentation: Manuscript preparation, peer review process, publication standards.",
            "Funding and Budget: Grants, funding sources, budget allocation.",
            "Project Timeline: Milestones, deliverables, deadlines.",
            "Gene expression analysis: Methods, tools, results interpretation.",
            "CRISPR-Cas9: Applications, ethical considerations, experimental protocols.",
            "Stem cell research: Differentiation, culture conditions, clinical applications.",
            "Clinical trial phases: Design, patient recruitment, outcome measures.",
            "Immunology: Antibody generation, vaccine development, immune response assays.",
            "Proteomics: Protein expression, purification, mass spectrometry.",
            "Genomics: DNA sequencing, genome annotation, variant analysis.",
            "Synthetic biology: Gene circuits, metabolic engineering, biosensors.",
            "Drug development: Lead optimization, pharmacokinetics, toxicology.",
            "Cell culture techniques: Primary cells, cell lines, bioreactors.",
            "Molecular biology: Cloning, PCR, gene editing.",
            "Microbiology: Bacterial culture, pathogen identification, antimicrobial resistance.",
            "Neuroscience: Brain imaging, neurophysiology, cognitive testing.",
            "Bioinformatics: Sequence analysis, structural modeling, data mining.",
            "Ecology: Population dynamics, ecosystem modeling, conservation biology.",
            "Zoology: Animal behavior, taxonomy, phylogenetics.",
            "Botany: Plant physiology, genetic modification, crop science.",
            "Evolutionary biology: Speciation, natural selection, phylogenetic trees.",
            "Biochemistry: Enzyme kinetics, metabolic pathways, molecular interactions.",
            "Biophysics: Protein folding, membrane dynamics, single-molecule techniques.",
            "Pharmacology: Drug-receptor interactions, dose-response relationships, pharmacodynamics.",
            "Biostatistics: Experimental design, data analysis, survival analysis.",
            "Tissue engineering: Scaffold design, stem cells, organ regeneration.",
            "Oncology: Cancer biomarkers, targeted therapy, clinical trials.",
            "Virology: Virus isolation, vaccine development, antiviral therapies.",
            "Plant biotechnology: Genetic engineering, tissue culture, crop improvement.",
            "Marine biology: Coral reefs, marine ecosystems, conservation strategies.",
            "Forensic biology: DNA fingerprinting, toxicology, crime scene analysis.",
            "Developmental biology: Embryogenesis, gene regulation, morphogenesis.",
            "Endocrinology: Hormone signaling, metabolic disorders, endocrine therapies.",
            "Biological imaging: Microscopy techniques, image analysis, fluorescent markers.",
            "Evolutionary genetics: Population genetics, molecular evolution, adaptive traits.",
            "Structural biology: Protein crystallography, NMR spectroscopy, electron microscopy.",
            "Metagenomics: Microbial community analysis, environmental samples, bioinformatics.",
            "Bioethics: Stem cell research, genetic testing, cloning ethics.",
            "Biomaterials: Nanoparticles, drug delivery systems, biodegradable polymers.",
            "Paleobiology: Fossil analysis, evolutionary history, extinct species.",
            "Ecotoxicology: Studying the effects of toxic chemicals on biological organisms.",
            "Genetic_Modification: Altering the genetic material of organisms for research or application.",
            "Biological_Therapeutics: Developing drugs and treatments derived from biological sources.",
            "Environmental_Biology: Studying the interactions between organisms and their environment.",
            "Epidemiology: Studying the distribution and determinants of health-related states in populations.",
            "Systems_Biology: Integrating biological data to understand complex biological systems.",
            "Bioengineering: Applying engineering principles to biological and medical problems.",
            "Neurobiology: Studying the structure and function of the nervous system.",
            "Immunogenetics: Studying the genetic basis of the immune response.",
            "Marine_Ecology: Studying marine organisms and their interactions with the environment.",
            # Common examples
            "Ensure compatibility with multiple operating systems, including Windows, macOS, and Linux.",
            "Implement features for real-time collaboration, allowing multiple users to edit and comment on documents simultaneously.",
            "Develop automated backup and recovery procedures to protect user data in the event of a system failure.",
            "Support both cloud-based and on-premise deployment options for flexibility and scalability.",
            "Design an intuitive user interface with customizable themes and accessibility options.",
            "Develop robust API endpoints for seamless integration with third-party applications and services."
        ]
        
        y_train = [
            # E-commerce examples
            "e-commerce", "e-commerce", "e-commerce", "e-commerce", "e-commerce", "e-commerce","e-commerce", "e-commerce", "e-commerce", "e-commerce",
            "e-commerce", "e-commerce", "e-commerce", "e-commerce", "e-commerce", "e-commerce","e-commerce", "e-commerce", "e-commerce", "e-commerce",
            "e-commerce", "e-commerce", "e-commerce", "e-commerce", "e-commerce", "e-commerce","e-commerce", "e-commerce", "e-commerce", "e-commerce",
            "e-commerce", "e-commerce", "e-commerce", "e-commerce", "e-commerce", "e-commerce","e-commerce", "e-commerce", "e-commerce", "e-commerce",
            "e-commerce", "e-commerce", "e-commerce", "e-commerce", "e-commerce", "e-commerce","e-commerce", "e-commerce", "e-commerce", "e-commerce",
            "e-commerce", "e-commerce", "e-commerce", "e-commerce", "e-commerce", "e-commerce","e-commerce", "e-commerce", "e-commerce", "e-commerce",
        
        
        
        
            # Financial examples
            "financial", "financial", "financial", "financial", "financial", "financial","financial", "financial", "financial", "financial",
            "financial", "financial", "financial", "financial", "financial", "financial","financial", "financial", "financial", "financial",
            "financial", "financial", "financial", "financial", "financial", "financial","financial", "financial", "financial", "financial",
            "financial", "financial", "financial", "financial", "financial", "financial","financial", "financial", "financial", "financial",
            "financial", "financial", "financial", "financial", "financial", "financial","financial", "financial", "financial", "financial",
            "financial", "financial", "financial", "financial", "financial", "financial","financial", "financial", "financial", "financial",
        
        
            # Biological examples
            "biological", "biological", "biological", "biological", "biological", "biological","biological", "biological", "biological", "biological",
            "biological", "biological", "biological", "biological", "biological", "biological","biological", "biological", "biological", "biological",
            "biological", "biological", "biological", "biological", "biological", "biological","biological", "biological", "biological", "biological",
            "biological", "biological", "biological", "biological", "biological", "biological","biological", "biological", "biological", "biological",
            "biological", "biological", "biological", "biological", "biological", "biological","biological", "biological", "biological", "biological",
            "biological", "biological", "biological", "biological", "biological", "biological","biological", "biological", "biological", "biological",
        
        
        
            # Common examples
            "common", "common", "common", "common", "common", "common"
        ]
        
        # Train the BERT classifier
        classifier, tokenizer_classifier, label_encoder = train_bert_classifier(X_train, y_train)
        
        
        # Select the appropriate model for the PRD
        selected_model, selected_tokenizer = select_model(prd_text, classifier, tokenizer_classifier, label_encoder)
        
        
        # Fine-tune the selected model using the PRD
        fine_tuned_model = fine_tune_model_with_prd(prd_text, selected_model, selected_tokenizer)
        
        
        # Refine the PRD using the selected model
        refined_prd_text = refine_prd_with_model(prd_text, fine_tuned_model, selected_tokenizer)
        # st.success("refined  successfully!")
        
        # Generate test cases using the Groq API
        # import pandas as pd

# Generate test cases
        test_cases_string = generate_test_cases(prd_text, api_key)
        st.success("TC Generated successfully!")
        st.write("Generated Test Cases:")
        st.write(test_cases_string)
        
        # # Parse test cases into a structured format (assume test cases are separated by line breaks)
        # test_cases_list = test_cases_string.split("\n")  # Modify the split logic if the format is different
        
        # # Create a DataFrame for the table
        # test_cases_df = pd.DataFrame({
        #     "Test Case ID": [f"TC-{i+1}" for i in range(len(test_cases_list))],
        #     "Description": test_cases_list
        # })
        
        # # Display the table
        # # st.write("Test Cases in Table Format:")
        # st.dataframe(test_cases_df)
        
        # # Save the DataFrame to a CSV file
        # output_csv_path = "generated_test_cases.csv"
        # test_cases_df.to_csv(output_csv_path, index=False)
        
        # # Allow the user to download the CSV file
        # with open(output_csv_path, "rb") as file:
        #     test_cases_csv_data = file.read()
        
        # st.download_button(
        #     label="Download",
        #     data=test_cases_csv_data,
        #     file_name="generated_test_cases.csv",
        #     mime="text/csv"
        # )

    else:
        st.info("Please select the option")
        

    

